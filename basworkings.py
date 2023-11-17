import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles
from io import BytesIO
import base64
import os
from datetime import datetime

combined_sheets_filename = "bas_workings.xlsx"
template_filename = "template.xlsx"


if os.path.exists(combined_sheets_filename):
    os.remove(combined_sheets_filename)


def copy_style(source_cell, dest_cell):
    if source_cell.value is not None:
        dest_cell.font = openpyxl.styles.Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            color=source_cell.font.color
        ) if source_cell.font is not None else None

        dest_cell.border = openpyxl.styles.Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        ) if source_cell.border is not None else None

        dest_cell.alignment = openpyxl.styles.Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        ) if source_cell.alignment is not None else None

        if isinstance(source_cell.value, (int, float)) and source_cell.number_format:
            dest_cell.number_format = source_cell.number_format

        if hasattr(source_cell, 'row') and (source_cell.row == 1 or (
                source_cell.fill is not None and source_cell.fill.start_color.index != "00000000")):
            dest_cell.alignment = openpyxl.styles.Alignment(
                vertical="center", wrap_text=source_cell.alignment.wrap_text
            )


def update_template_summary(template_wb, start_date, end_date, selected_sheets):
    template_sheet = template_wb["Summary"]

    for merged_range in template_sheet.merged_cells.ranges:
        if 7 <= merged_range.min_row <= 8 and 1 <= merged_range.min_col <= 13:
            merged_cell = template_sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            merged_cell.value = f"For the period {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"
            merged_cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Find the last column index for the sheet names
    last_col_idx = 3

    # Add sheet names in bold and centered in columns D, E, F, ...
    for file_info in selected_sheets.values():
        sheet_name = file_info["sheet_name"]

        # Skip "Intra-GST transactions"
        if sheet_name == "Intra-GST transactions":
            continue

        last_col_idx += 1
        sheet_name_cell = template_sheet.cell(row=9, column=last_col_idx)
        sheet_name_cell.value = sheet_name
        sheet_name_cell.font = openpyxl.styles.Font(bold=True)
        sheet_name_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        col_letter = openpyxl.utils.get_column_letter(last_col_idx)

        col_letter2 = openpyxl.utils.get_column_letter(last_col_idx+3)
        col_letter3 = openpyxl.utils.get_column_letter(last_col_idx+4)

        # Calculate the sum and update D10
        sum_formula = f"SUM('{sheet_name}'!C7:D7)"
        template_sheet.cell(row=10, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C8:D8)"
        template_sheet.cell(row=11, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C9:D9)"
        template_sheet.cell(row=12, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C10:D10)"
        template_sheet.cell(row=13, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C11:D11)"
        template_sheet.cell(row=14, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C12:D12)"
        template_sheet.cell(row=15, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C13:D13)"
        template_sheet.cell(row=16, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C14:D14)"
        template_sheet.cell(row=17, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C15:D15)"
        template_sheet.cell(row=18, column=last_col_idx).value = f"={sum_formula}"


        sum_formula = f"SUM('{sheet_name}'!C18:D18)"
        template_sheet.cell(row=21, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C19:D19)"
        template_sheet.cell(row=22, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C20:D20)"
        template_sheet.cell(row=23, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C21:D21)"
        template_sheet.cell(row=24, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C22:D22)"
        template_sheet.cell(row=25, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C23:D23)"
        template_sheet.cell(row=26, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C24:D24)"
        template_sheet.cell(row=27, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C25:D25)"
        template_sheet.cell(row=28, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C26:D26)"
        template_sheet.cell(row=29, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C27:D27)"
        template_sheet.cell(row=30, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"SUM('{sheet_name}'!C28:D28)"
        template_sheet.cell(row=31, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"{col_letter}18-{col_letter}31"
        template_sheet.cell(row=34, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"sum({col_letter}34:{col_letter}41)"
        template_sheet.cell(row=43, column=last_col_idx).value = f"={sum_formula}"

        sum_formula = f"sum(C34:C41)"
        template_sheet.cell(row=43, column=3).value = f"={sum_formula}"

        sum_formula = f"sum({col_letter2}34:{col_letter2}41)"
        template_sheet.cell(row=43, column=last_col_idx + 3).value = f"={sum_formula}"
        template_sheet.cell(row=43, column=last_col_idx + 3).font = openpyxl.styles.Font(bold=True)
        template_sheet.cell(row=43, column=last_col_idx + 1).font = openpyxl.styles.Font(bold=False)
        template_sheet.cell(row=43, column=last_col_idx + 2).font = openpyxl.styles.Font(bold=False)

        sum_formula = f"sum({col_letter3}34:{col_letter3}41)"
        template_sheet.cell(row=43, column=last_col_idx + 4).value = f"={sum_formula}"

    if last_col_idx > 3:
        total_col_letter = openpyxl.utils.get_column_letter(last_col_idx)
        for row_idx in range(10, 32):
            total_formula = f"SUM(D{row_idx}:{total_col_letter}{row_idx})"
            template_sheet.cell(row=row_idx, column=last_col_idx + 1).value = f"={total_formula}"

        # Calculate and update "Adjustment" column in row 10
        adjustment_formula = (
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST on Income", \'Intra-GST transactions\'!G:G) + '
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST Free Income", \'Intra-GST transactions\'!G:G)'
        )
        template_sheet.cell(row=10, column=last_col_idx + 2).value = f"={adjustment_formula}"

        adjustment_formula = (
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST Free Income", \'Intra-GST transactions\'!G:G)'
        )
        template_sheet.cell(row=12, column=last_col_idx + 2).value = f"={adjustment_formula}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}11 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}12 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}13'
        )
        template_sheet.cell(row=14, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}10 - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}14'
        )
        template_sheet.cell(row=15, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}15 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}16'
        )
        template_sheet.cell(row=17, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}17/11 '
        )
        template_sheet.cell(row=18, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula = (
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST on Expenses", \'Intra-GST transactions\'!G:G) + '
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST Free Expenses", \'Intra-GST transactions\'!G:G)'
        )
        template_sheet.cell(row=22, column=last_col_idx + 2).value = f"={adjustment_formula}"

        adjustment_formula = (
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST Free Expenses", \'Intra-GST transactions\'!G:G)'
        )
        template_sheet.cell(row=25, column=last_col_idx + 2).value = f"={adjustment_formula}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}21 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}22'
        )
        template_sheet.cell(row=23, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}24 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}25 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}26'
        )
        template_sheet.cell(row=27, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}23 - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}27'
        )
        template_sheet.cell(row=28, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula_row = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}28 + '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}29'
        )
        template_sheet.cell(row=30, column=last_col_idx + 2).value = f"={adjustment_formula_row}"

        adjustment_formula = (
            f'SUMIF(\'Intra-GST transactions\'!B:B, "GST Free Expenses", \'Intra-GST transactions\'!H:H)'
        )
        template_sheet.cell(row=31, column=last_col_idx + 2).value = f"={adjustment_formula}"


    for row_idx in range(10, 32):
        total_after_adjustment_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 1)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 3).value = f"={total_after_adjustment_formula}"

        checking_formula = f"={openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}"
        template_sheet.cell(row=row_idx, column=last_col_idx + 4).value = checking_formula

        adj_balance = f"={openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}"
        template_sheet.cell(row=row_idx, column=3).value = f"={adj_balance}"


        adj_balance = f"={openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}"
        template_sheet.cell(row=row_idx, column=3).value = f"={adj_balance}"

        blank_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 5).value = f"={blank_formula}"

    for row_idx in range(10, 11):
        total_after_adjustment_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}10 - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}18'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 7).value = f"={total_after_adjustment_formula}"
        template_sheet.cell(row=row_idx, column=last_col_idx + 7).number_format = '#,##0_);[Black](#,##0)'

    for row_idx in range(34, 35):
        adj_balance = f"={openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}"
        template_sheet.cell(row=row_idx, column=3).value = f"={adj_balance}"

        total_after_adjustment_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}18 - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}31'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 3).value = f"={total_after_adjustment_formula}"

        checking_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}18 - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}31'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 4).value = f"={checking_formula}"

        total_formula = f"SUM(D{row_idx}:{total_col_letter}{row_idx})"
        template_sheet.cell(row=row_idx, column=last_col_idx + 1).value = f"={total_formula}"

        blank_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 5).value = f"={blank_formula}"

    for row_idx in range(37, 42):
        adj_balance = f"={openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}"
        template_sheet.cell(row=row_idx, column=3).value = f"={adj_balance}"

        total_after_adjustment_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 1)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 3).value = f"={total_after_adjustment_formula}"

        checking_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 2)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 4).value = f"={checking_formula}"

        total_formula = f"SUM(D{row_idx}:{total_col_letter}{row_idx})"
        template_sheet.cell(row=row_idx, column=last_col_idx + 1).value = f"={total_formula}"

        blank_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 5).value = f"={blank_formula}"

    for row_idx in range(42, 43):
        blank_formula = (
            f'{openpyxl.utils.get_column_letter(last_col_idx + 3)}{row_idx} - '
            f'{openpyxl.utils.get_column_letter(last_col_idx + 4)}{row_idx}'
        )
        template_sheet.cell(row=row_idx, column=last_col_idx + 5).value = f"={blank_formula}"


    for col_idx in range(3, last_col_idx + 5):
        for row_idx in range(7, template_sheet.max_row + 1):
            cell = template_sheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '#,##0_);[Black](#,##0)'
            if 9 <= row_idx <= 43 and col_idx == last_col_idx + 3:
                cell.fill = openpyxl.styles.PatternFill(start_color="DEDEDE", end_color="DEDEDE", fill_type="solid")

                

    # Add "Total", "Adjustment", "Total after adjustment", and "Checking"
    total_cell = template_sheet.cell(row=9, column=last_col_idx + 1)
    total_cell.value = "Total"
    total_cell.font = openpyxl.styles.Font(bold=True)
    total_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    adjustment_cell = template_sheet.cell(row=9, column=last_col_idx + 2)
    adjustment_cell.value = "Adjustment"
    adjustment_cell.font = openpyxl.styles.Font(bold=True)
    adjustment_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    total_after_adjustment_cell = template_sheet.cell(row=9, column=last_col_idx + 3)
    total_after_adjustment_cell.value = "Total after adjustment"
    total_after_adjustment_cell.font = openpyxl.styles.Font(bold=True)
    total_after_adjustment_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    checking_cell = template_sheet.cell(row=9, column=last_col_idx + 4)
    checking_cell.value = "Checking"
    checking_cell.font = openpyxl.styles.Font(bold=True)
    checking_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")


    total_col_idx = last_col_idx + 1
    adjustment_col_idx = last_col_idx + 2
    total_after_adjustment_col_idx = last_col_idx + 3
    checking_col_idx = last_col_idx + 4

    # Apply accounting format with no decimals to cells in columns C to N
    for col_idx in range(3, last_col_idx + 5):
        for row_idx in range(7, template_sheet.max_row + 1):
            cell = template_sheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '#,##0_);[Black](#,##0)'

    return total_col_idx, adjustment_col_idx, total_after_adjustment_col_idx, checking_col_idx


def copy_sheet_to_main_workbook(file, sheet_name, main_wb):
    source_wb = openpyxl.load_workbook(file, read_only=False, data_only=True)
    source_sheet = source_wb[sheet_name]

    # Create a new sheet in the main workbook
    main_sheet = main_wb.create_sheet(sheet_name)


    for row_idx, row in enumerate(dataframe_to_rows(pd.DataFrame(source_sheet.values), index=False, header=True), 1):
        if row_idx > 1:  # Skip the first row
            for col_idx, value in enumerate(row, 1):
                dest_cell = main_sheet.cell(row=row_idx - 1, column=col_idx, value=value)

                # Copy styles
                source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                copy_style(source_cell, dest_cell)
                
                # Copy formula
                if source_cell.data_type == 'f':
                    dest_cell.value = source_cell.value



    if sheet_name not in ["Summary", "Intra-GST transactions"]:
        for row_idx in range(1, 4):
            main_sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)

        # Make rows 1-4 bold and centered

        for row_idx in range(1, 5):
            for col_idx in range(1, 9):
                cell = main_sheet.cell(row=row_idx, column=col_idx)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")




    main_wb.save(combined_sheets_filename)


def combine_sheets(selected_sheets, start_date, end_date):
    if os.path.exists(combined_sheets_filename):
        main_wb = load_workbook(combined_sheets_filename)
    else:
        main_wb = load_workbook(template_filename)
        main_wb.save(combined_sheets_filename)

    # Update the Summary sheet in the template
    update_template_summary(main_wb, start_date, end_date, selected_sheets)

    for file_info in selected_sheets.values():
        file = file_info["file"]
        sheet_name = file_info["sheet_name"]

        copy_sheet_to_main_workbook(file, sheet_name, main_wb)

    main_wb.close()


def download_link():
    combined_wb = load_workbook(combined_sheets_filename)
    towrite = BytesIO()
    combined_wb.save(towrite)
    towrite.seek(0)
    b64 = base64.b64encode(towrite.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="bas_workings.xlsx">Download Combined Sheets</a>'
    st.markdown(href, unsafe_allow_html=True)

# Streamlit app
st.title("BAS Workings Consolidation App")

uploaded_files = st.file_uploader("Upload Excel Files", type=["xlsx"], accept_multiple_files=True)

selected_sheets = {}
start_date = st.date_input("Select Start Date", datetime.today())
end_date = st.date_input("Select End Date", datetime.today())

if uploaded_files:
    for file in uploaded_files:
        st.subheader(f"File: {file.name}")
        sheet_names = pd.ExcelFile(file).sheet_names
        sheet_name = st.selectbox("Select a sheet:", sheet_names)
        selected_sheets[file.name] = {"file": file, "sheet_name": sheet_name}

    if st.button("Combine Sheets"):
        combine_sheets(selected_sheets, start_date, end_date)
        st.success("Sheets combined successfully!")
        download_link()
